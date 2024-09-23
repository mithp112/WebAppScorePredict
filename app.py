from flask import Flask, render_template, request, send_file
import pandas as pd
import joblib
import numpy as np
from concurrent.futures import ThreadPoolExecutor
from tensorflow.keras.models import load_model, Model
from tensorflow.keras.layers import Flatten, Dense, LSTM
from tensorflow.keras.saving import register_keras_serializable
from openpyxl.utils import get_column_letter
import io, json
import sqlite3
from datetime import date

@register_keras_serializable()
class MLPModel1(Model):
    def __init__(self):
        super(MLPModel1, self).__init__()
        self.flatten = Flatten()
        self.fc1 = Dense(128, activation='relu')
        self.fc2 = Dense(62, activation='relu') 
        self.fc3 = Dense(32, activation='relu')
        self.fc4 = Dense(18, activation='linear') 
    def call(self, inputs):
        features = self.flatten(inputs)
        features = self.fc1(features)
        features = self.fc2(features)
        features = self.fc3(features)
        features = self.fc4(features)
        return features
    
    def get_config(self):
        config = super(MLPModel1, self).get_config()
        # Return the config dictionary including any custom parameters
        return config

    @classmethod
    def from_config(cls, config):
        # Create a new instance of the model from the config
        return cls()
    
@register_keras_serializable()
class MLPModel2(Model):
    def __init__(self):
        super(MLPModel2, self).__init__()
        self.flatten = Flatten()
        self.fc1 = Dense(128, activation='relu')
        self.fc2 = Dense(62, activation='relu') 
        self.fc3 = Dense(32, activation='relu')
        self.fc4 = Dense(6, activation='linear') 
    def call(self, inputs):
        features = self.flatten(inputs)
        features = self.fc1(features)
        features = self.fc2(features)
        features = self.fc3(features)
        features = self.fc4(features)
        return features
    
    def get_config(self):
        config = super(MLPModel1, self).get_config()
        # Return the config dictionary including any custom parameters
        return config

    @classmethod
    def from_config(cls, config):
        # Create a new instance of the model from the config
        return cls()

@register_keras_serializable()
class LSTMModel1(Model):
    def __init__(self):
        super(LSTMModel1, self).__init__()
        self.LSTM1 = LSTM(128, activation = 'relu', return_sequences = True)
        self.LSTM2 = LSTM(64, activation = 'relu', return_sequences= True)
        self.LSTM3 = LSTM(32, activation = 'relu', return_sequences= False)
        self.fc1 = Dense(18, activation = 'linear')


    def call(self, inputs):
        features = self.LSTM1(inputs)
        features = self.LSTM2(features)
        features = self.LSTM3(features)
        features = self.fc1(features)
        return features
  
    def get_config(self):
        config = super(LSTMModel1, self).get_config()
        # Return the config dictionary including any custom parameters
        return config

    @classmethod
    def from_config(cls, config):
        # Create a new instance of the model from the config
        return cls()
    
@register_keras_serializable()
class LSTMModel2(Model):
    def __init__(self):
        super(LSTMModel2, self).__init__()
        self.LSTM1 = LSTM(128, activation = 'relu', return_sequences = True)
        self.LSTM2 = LSTM(64, activation = 'relu', return_sequences= True)
        self.LSTM3 = LSTM(32, activation = 'relu', return_sequences= False)
        self.fc1 = Dense(6, activation = 'linear')


    def call(self, inputs):
        features = self.LSTM1(inputs)
        features = self.LSTM2(features)
        features = self.LSTM3(features)
        features = self.fc1(features)
        return features
  
    def get_config(self):
        config = super(LSTMModel1, self).get_config()
        # Return the config dictionary including any custom parameters
        return config

    @classmethod
    def from_config(cls, config):
        # Create a new instance of the model from the config
        return cls()


app = Flask(__name__, template_folder='templates')

@app.route('/')   
def home():
    return render_template('Page1.html')

@app.route('/page1')
def page1():
    return render_template('Page1.html')

@app.route('/page2')
def page2():
    return render_template('Page2.html')

def predModel1(ypred): # hiển thị kết quả cho model 1
    pred = pd.DataFrame({
        'Toán_1': pd.Series(ypred[:,0]),
        'Văn_1': pd.Series(ypred[:,1]),
        'Lý_1': pd.Series(ypred[:,2]),
        'Hóa_1': pd.Series(ypred[:,3]),
        'Sinh_1': pd.Series(ypred[:,4]),
        'Sử_1': pd.Series(ypred[:,5]),
        'Địa_1': pd.Series(ypred[:,6]),
        'Anh_1': pd.Series(ypred[:,7]),
        'GDCD_1': pd.Series(ypred[:,8]),
        'Toán_2': pd.Series(ypred[:,9]),
        'Văn_2': pd.Series(ypred[:,10]),
        'Lý_2': pd.Series(ypred[:,11]),
        'Hóa_2': pd.Series(ypred[:,12]),
        'Sinh_2': pd.Series(ypred[:,13]),
        'Sử_2': pd.Series(ypred[:,14]),
        'Địa_2': pd.Series(ypred[:,15]),
        'Anh_2': pd.Series(ypred[:,16]),
        'GDCD_2': pd.Series(ypred[:,17])
    })
    pred = pred.astype('float64')
    return pred

def predModel2(ypred): # hiển thị kết quả cho model 2
    pred = pd.DataFrame({
        'Toán_1': pd.Series(ypred[:,0]),
        'Văn_1': pd.Series(ypred[:,1]),
        'Lý_1': pd.Series(ypred[:,2]),
        'Hóa_1': pd.Series(ypred[:,3]),
        'Sinh_1': pd.Series(ypred[:,4]),
        'Anh_1': pd.Series(ypred[:,5]),
    })
    pred = pred.astype('float64')
    return pred


    
@app.route("/predict1", methods = ['GET','POST'])
def predict1():
    log_access('/predict1')
    pred_df1 = None
    pred_df2 = None
    pred_df3 = None
    if request.method == "POST":
        try:
            Toan_1_10 = float(request.form['Toan_1_10'])
            Toan_2_10 = float(request.form['Toan_2_10'])
            Toan_1_11 = float(request.form['Toan_1_11'])
            Toan_2_11 = float(request.form['Toan_2_11'])

            Van_1_10 = float(request.form['Van_1_10'])
            Van_2_10 = float(request.form['Van_2_10'])
            Van_1_11 = float(request.form['Van_1_11'])
            Van_2_11 = float(request.form['Van_2_11'])

            Ly_1_10 = float(request.form['Ly_1_10'])
            Ly_2_10 = float(request.form['Ly_2_10'])
            Ly_1_11 = float(request.form['Ly_1_11'])
            Ly_2_11 = float(request.form['Ly_2_11'])

            Anh_1_10 = float(request.form['Anh_1_10'])
            Anh_2_10 = float(request.form['Anh_2_10'])
            Anh_1_11 = float(request.form['Anh_1_11'])
            Anh_2_11 = float(request.form['Anh_2_11'])

            Su_1_10 = float(request.form['Su_1_10'])
            Su_2_10 = float(request.form['Su_2_10'])
            Su_1_11 = float(request.form['Su_1_11'])
            Su_2_11 = float(request.form['Su_2_11'])

            Dia_1_10 = float(request.form['Dia_1_10'])
            Dia_2_10 = float(request.form['Dia_2_10'])
            Dia_1_11 = float(request.form['Dia_1_11'])
            Dia_2_11 = float(request.form['Dia_2_11'])

            Sinh_1_10 = float(request.form['Sinh_1_10'])
            Sinh_2_10 = float(request.form['Sinh_2_10'])
            Sinh_1_11 = float(request.form['Sinh_1_11'])
            Sinh_2_11 = float(request.form['Sinh_2_11'])

            Hoa_1_10 = float(request.form['Hoa_1_10'])
            Hoa_2_10 = float(request.form['Hoa_2_10'])
            Hoa_1_11 = float(request.form['Hoa_1_11'])
            Hoa_2_11 = float(request.form['Hoa_2_11'])

            GDCD_1_10 = float(request.form['GDCD_1_10'])
            GDCD_2_10 = float(request.form['GDCD_2_10'])
            GDCD_1_11 = float(request.form['GDCD_1_11'])
            GDCD_2_11 = float(request.form['GDCD_2_11'])
            orphan = 1 if 'orphan' in request.form else 0
            kios = 2 if 'kios' in request.form else 0
            orphan_and_kios = kios + orphan

            pred_arg = [Toan_1_10,Van_1_10,Ly_1_10,Hoa_1_10,Sinh_1_10,Su_1_10,Dia_1_10,Anh_1_10, GDCD_1_10
                        ,Toan_2_10,Van_2_10,Ly_2_10,Hoa_2_10,Sinh_2_10,Su_2_10,Dia_2_10,Anh_2_10, GDCD_2_10
                        ,Toan_1_11,Van_1_11,Ly_1_11,Hoa_1_11,Sinh_1_11,Su_1_11,Dia_1_11,Anh_1_11, GDCD_1_11
                        ,Toan_2_11,Van_2_11,Ly_2_11,Hoa_2_11,Sinh_2_11,Su_2_11,Dia_2_11,Anh_2_11, GDCD_2_11
                        ,orphan_and_kios]
            
            pred_arg_arr = np.array(pred_arg)
            pred_arg_arr = pred_arg_arr.reshape(1,-1)
            def load_and_predict(model_path, model_type):
                if model_type == 'keras':
                    model = load_model(f"Models/{model_path}")  # Load model Keras
                else:
                    with open(f"Models/{model_path}", 'rb') as model_file:
                        model = joblib.load(model_file)  # Load model với joblib

                if model_type == 'keras' and "LSTM" in model_path:  # Kiểm tra xem đây có phải là model LSTM không
                    return model.predict(pred_arg_arr.reshape(pred_arg_arr.shape[0], 1, pred_arg_arr.shape[1]))
                else:
                    return model.predict(pred_arg_arr)
            with ThreadPoolExecutor() as executor:
                future1 = executor.submit(load_and_predict, "LR_10_11_12.pkl", model_type='joblib')
                future2 = executor.submit(load_and_predict, "MLP_10_11_12.keras", model_type='keras')
                future3 = executor.submit(load_and_predict, "LSTM_10_11_12.keras", model_type='keras')

                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel1(prediction1)
                pred_df2 = predModel1(prediction2)
                pred_df3 = predModel1(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))
                

        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
    return render_template('Page3.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
)

@app.route("/predict2", methods = ['GET','POST'])
def predict2():
    log_access('/predict2')
    if request.method == "POST":
        try:
            Toan_1_10 = float(request.form['Toan_1_10'])
            Toan_2_10 = float(request.form['Toan_2_10'])
            Toan_1_11 = float(request.form['Toan_1_11'])
            Toan_2_11 = float(request.form['Toan_2_11'])
            Toan_1_12 = float(request.form['Toan_1_12'])
            Toan_2_12 = float(request.form['Toan_2_12'])

            Van_1_10 = float(request.form['Van_1_10'])
            Van_2_10 = float(request.form['Van_2_10'])
            Van_1_11 = float(request.form['Van_1_11'])
            Van_2_11 = float(request.form['Van_2_11'])
            Van_1_12 = float(request.form['Van_1_12'])
            Van_2_12 = float(request.form['Van_2_12'])

            Ly_1_10 = float(request.form['Ly_1_10'])
            Ly_2_10 = float(request.form['Ly_2_10'])
            Ly_1_11 = float(request.form['Ly_1_11'])
            Ly_2_11 = float(request.form['Ly_2_11'])
            Ly_1_12 = float(request.form['Ly_1_12'])
            Ly_2_12 = float(request.form['Ly_2_12'])

            Anh_1_10 = float(request.form['Anh_1_10'])
            Anh_2_10 = float(request.form['Anh_2_10'])
            Anh_1_11 = float(request.form['Anh_1_11'])
            Anh_2_11 = float(request.form['Anh_2_11'])
            Anh_1_12 = float(request.form['Anh_1_12'])
            Anh_2_12 = float(request.form['Anh_2_12'])

            Su_1_10 = float(request.form['Su_1_10'])
            Su_2_10 = float(request.form['Su_2_10'])
            Su_1_11 = float(request.form['Su_1_11'])
            Su_2_11 = float(request.form['Su_2_11'])
            Su_1_12 = float(request.form['Su_1_12'])
            Su_2_12 = float(request.form['Su_2_12'])


            Dia_1_10 = float(request.form['Dia_1_10'])
            Dia_2_10 = float(request.form['Dia_2_10'])
            Dia_1_11 = float(request.form['Dia_1_11'])
            Dia_2_11 = float(request.form['Dia_2_11'])
            Dia_1_12 = float(request.form['Dia_1_12'])
            Dia_2_12 = float(request.form['Dia_2_12'])

            Sinh_1_10 = float(request.form['Sinh_1_10'])
            Sinh_2_10 = float(request.form['Sinh_2_10'])
            Sinh_1_11 = float(request.form['Sinh_1_11'])
            Sinh_2_11 = float(request.form['Sinh_2_11'])
            Sinh_1_12 = float(request.form['Sinh_1_12'])
            Sinh_2_12 = float(request.form['Sinh_2_12'])

            Hoa_1_10 = float(request.form['Hoa_1_10'])
            Hoa_2_10 = float(request.form['Hoa_2_10'])
            Hoa_1_11 = float(request.form['Hoa_1_11'])
            Hoa_2_11 = float(request.form['Hoa_2_11'])
            Hoa_1_12 = float(request.form['Hoa_1_12'])
            Hoa_2_12 = float(request.form['Hoa_2_12'])

            GDCD_1_10 = float(request.form['GDCD_1_10'])
            GDCD_2_10 = float(request.form['GDCD_2_10'])
            GDCD_1_11 = float(request.form['GDCD_1_11'])
            GDCD_2_11 = float(request.form['GDCD_2_11'])
            GDCD_1_12 = float(request.form['GDCD_1_12'])
            GDCD_2_12 = float(request.form['GDCD_2_12'])
            orphan = 1 if 'orphan' in request.form else 0
            kios = 2 if 'kios' in request.form else 0
            orphan_and_kios = kios + orphan
            

            pred_arg = [Toan_1_10,Van_1_10,Ly_1_10,Hoa_1_10,Sinh_1_10,Su_1_10,Dia_1_10,Anh_1_10, GDCD_1_10
                        ,Toan_2_10,Van_2_10,Ly_2_10,Hoa_2_10,Sinh_2_10,Su_2_10,Dia_2_10,Anh_2_10, GDCD_2_10
                        ,Toan_1_11,Van_1_11,Ly_1_11,Hoa_1_11,Sinh_1_11,Su_1_11,Dia_1_11,Anh_1_11, GDCD_1_11
                        ,Toan_2_11,Van_2_11,Ly_2_11,Hoa_2_11,Sinh_2_11,Su_2_11,Dia_2_11,Anh_2_11, GDCD_2_11
                        ,Toan_1_12,Van_1_12,Ly_1_12,Hoa_1_12,Sinh_1_12,Su_1_12,Dia_1_12,Anh_1_12, GDCD_1_12
                        ,Toan_2_12,Van_2_12,Ly_2_12,Hoa_2_12,Sinh_2_12,Su_2_12,Dia_2_12,Anh_2_12, GDCD_2_12
                        ,orphan_and_kios]
            
            
            pred_arg_arr = np.array(pred_arg)
            pred_arg_arr = pred_arg_arr.reshape(1,-1)
            def load_and_predict(model_path, model_type):
                if model_type == 'keras':
                    model = load_model(f"Models/{model_path}")  # Load model Keras
                else:
                    with open(f"Models/{model_path}", 'rb') as model_file:
                        model = joblib.load(model_file)  # Load model với joblib

                if model_type == 'keras' and "LSTM" in model_path:  # Kiểm tra xem đây có phải là model LSTM không
                    return model.predict(pred_arg_arr.reshape(pred_arg_arr.shape[0], 1, pred_arg_arr.shape[1]))
                else:
                    return model.predict(pred_arg_arr)
            nature_checked = True if 'nature' in request.form else False
            subjects = {}

            with ThreadPoolExecutor() as executor:
                if nature_checked:
                    print("call TN")
                    future1 = executor.submit(load_and_predict, "LR_TN_TN.pkl", model_type='joblib')
                    future2 = executor.submit(load_and_predict, "MLP_TN_TN.keras", model_type='keras')
                    future3 = executor.submit(load_and_predict, "LSTM_TN_TN.keras", model_type='keras')
                    subjects = {'mon1': 'Điểm Lý', 'mon1_1': 'Physics', 'mon2': 'Điểm Hóa', 'mon2_1': 'Chemistry', 'mon3': 'Điểm Sinh', 'mon3_1':  'Biology', 'type':  'TN'}
                else:
                    print("call XH")
                    future1 = executor.submit(load_and_predict, "LR_TN_XH.pkl", model_type='joblib')
                    future2 = executor.submit(load_and_predict, "MLP_TN_XH.keras", model_type='keras')
                    future3 = executor.submit(load_and_predict, "LSTM_TN_XH.keras", model_type='keras')
                    subjects = {'mon1': 'Điểm Sử', 'mon1_1': 'History', 'mon2': 'Điểm Địa', 'mon2_1': 'Geography', 'mon3': 'Điểm GDCD', 'mon3_1':  'Civic Education', 'type':  'XH'}
                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel2(prediction1)
                pred_df2 = predModel2(prediction2)
                pred_df3 = predModel2(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))
                
        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
    return render_template('Page4.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
    subjects=subjects
)



@app.route("/predict_excel1", methods=['GET', 'POST'])
def predict_excel1():
    log_access('/predict_excel1')
    if request.method == "POST":
        try:
            excel_file = request.files['excel_file']
            df = pd.read_excel(excel_file, header=None)  
            names = df.iloc[:, 0].tolist()
            df = df.drop(columns=0) 
            df = df.iloc[:, :37]
            pred_arg_arr = df.to_numpy()
            def load_and_predict(model_path, model_type):
                if model_type == 'keras':
                    model = load_model(f"Models/{model_path}")  # Load model Keras
                else:
                    with open(f"Models/{model_path}", 'rb') as model_file:
                        model = joblib.load(model_file)  # Load model với joblib

                if model_type == 'keras' and "LSTM" in model_path:  # Kiểm tra xem đây có phải là model LSTM không
                    return model.predict(pred_arg_arr.reshape(pred_arg_arr.shape[0], 1, pred_arg_arr.shape[1]))
                else:
                    return model.predict(pred_arg_arr)

           
            with ThreadPoolExecutor() as executor:
                future1 = executor.submit(load_and_predict, "LR_10_11_12.pkl", model_type='joblib')
                future2 = executor.submit(load_and_predict, "MLP_10_11_12.keras", model_type='keras')
                future3 = executor.submit(load_and_predict, "LSTM_10_11_12.keras", model_type='keras')

                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel1(prediction1)
                pred_df2 = predModel1(prediction2)
                pred_df3 = predModel1(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))
                
        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
            return {"error": str(e)}

    return render_template('Page3.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
    names=names
)


@app.route('/download_excel1', methods=['POST'])
def download_excel1():
    names = request.form.getlist('names')
    kq_Model1 = request.form.getlist('kq_Model1')
    kq_Model2 = request.form.getlist('kq_Model2')
    kq_Model3 = request.form.getlist('kq_Model3')

    names = json.loads(names[0].replace("'", '"'))
    kq_Model1 = json.loads(kq_Model1[0].replace("'", '"'))
    kq_Model2 = json.loads(kq_Model2[0].replace("'", '"'))
    kq_Model3 = json.loads(kq_Model3[0].replace("'", '"'))

    dataLR = {
        "Tên": names,

        # Kết quả từ LR model
        "Điểm Toán I": [kq_Model1['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn I": [kq_Model1['Văn_1'][i] for i in range(len(names))],
        "Điểm Lý I": [kq_Model1['Lý_1'][i] for i in range(len(names))],
        "Điểm Hóa I": [kq_Model1['Hóa_1'][i] for i in range(len(names))],
        "Điểm Sinh I": [kq_Model1['Sinh_1'][i] for i in range(len(names))],
        "Điểm Sử I": [kq_Model1['Sử_1'][i] for i in range(len(names))],
        "Điểm Địa I": [kq_Model1['Địa_1'][i] for i in range(len(names))],
        "Điểm GDCD I": [kq_Model1['GDCD_1'][i] for i in range(len(names))],
        "Điểm Anh I": [kq_Model1['Anh_1'][i] for i in range(len(names))],

        "Điểm Toán II": [kq_Model1['Toán_2'][i] for i in range(len(names))],
        "Điểm Văn II": [kq_Model1['Văn_2'][i] for i in range(len(names))],
        "Điểm Lý II": [kq_Model1['Lý_2'][i] for i in range(len(names))],
        "Điểm Hóa II": [kq_Model1['Hóa_2'][i] for i in range(len(names))],
        "Điểm Sinh II": [kq_Model1['Sinh_2'][i] for i in range(len(names))],
        "Điểm Sử II": [kq_Model1['Sử_2'][i] for i in range(len(names))],
        "Điểm Địa II": [kq_Model1['Địa_2'][i] for i in range(len(names))],
        "Điểm GDCD II": [kq_Model1['GDCD_2'][i] for i in range(len(names))],
        "Điểm Anh II": [kq_Model1['Anh_2'][i] for i in range(len(names))],
    }

    dataMLP = {
        "Tên": names,
        
        # Kết quả từ MLP model
        "Điểm Toán I": [kq_Model2['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn I": [kq_Model2['Văn_1'][i] for i in range(len(names))],
        "Điểm Lý I": [kq_Model2['Lý_1'][i] for i in range(len(names))],
        "Điểm Hóa I": [kq_Model2['Hóa_1'][i] for i in range(len(names))],
        "Điểm Sinh I": [kq_Model2['Sinh_1'][i] for i in range(len(names))],
        "Điểm Sử I": [kq_Model2['Sử_1'][i] for i in range(len(names))],
        "Điểm Địa I": [kq_Model2['Địa_1'][i] for i in range(len(names))],
        "Điểm GDCD I": [kq_Model2['GDCD_1'][i] for i in range(len(names))],
        "Điểm Anh I": [kq_Model2['Anh_1'][i] for i in range(len(names))],

        "Điểm Toán II": [kq_Model2['Toán_2'][i] for i in range(len(names))],
        "Điểm Văn II": [kq_Model2['Văn_2'][i] for i in range(len(names))],
        "Điểm Lý II": [kq_Model2['Lý_2'][i] for i in range(len(names))],
        "Điểm Hóa II": [kq_Model2['Hóa_2'][i] for i in range(len(names))],
        "Điểm Sinh II": [kq_Model2['Sinh_2'][i] for i in range(len(names))],
        "Điểm Sử II": [kq_Model2['Sử_2'][i] for i in range(len(names))],
        "Điểm Địa II": [kq_Model2['Địa_2'][i] for i in range(len(names))],
        "Điểm GDCD II": [kq_Model2['GDCD_2'][i] for i in range(len(names))],
        "Điểm Anh II": [kq_Model2['Anh_2'][i] for i in range(len(names))],
    }

    dataLSTM = {
        "Tên": names,

        # Kết quả từ LSTM model
        "Điểm Toán I": [kq_Model3['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn I": [kq_Model3['Văn_1'][i] for i in range(len(names))],
        "Điểm Lý I": [kq_Model3['Lý_1'][i] for i in range(len(names))],
        "Điểm Hóa I": [kq_Model3['Hóa_1'][i] for i in range(len(names))],
        "Điểm Sinh I": [kq_Model3['Sinh_1'][i] for i in range(len(names))],
        "Điểm Sử I": [kq_Model3['Sử_1'][i] for i in range(len(names))],
        "Điểm Địa I": [kq_Model3['Địa_1'][i] for i in range(len(names))],
        "Điểm GDCD I": [kq_Model3['GDCD_1'][i] for i in range(len(names))],
        "Điểm Anh I": [kq_Model3['Anh_1'][i] for i in range(len(names))],

        "Điểm Toán II": [kq_Model3['Toán_2'][i] for i in range(len(names))],
        "Điểm Văn II": [kq_Model3['Văn_2'][i] for i in range(len(names))],
        "Điểm Lý II": [kq_Model3['Lý_2'][i] for i in range(len(names))],
        "Điểm Hóa II": [kq_Model3['Hóa_2'][i] for i in range(len(names))],
        "Điểm Sinh II": [kq_Model3['Sinh_2'][i] for i in range(len(names))],
        "Điểm Sử II": [kq_Model3['Sử_2'][i] for i in range(len(names))],
        "Điểm Địa II": [kq_Model3['Địa_2'][i] for i in range(len(names))],
        "Điểm GDCD II": [kq_Model3['GDCD_2'][i] for i in range(len(names))],
        "Điểm Anh II": [kq_Model3['Anh_2'][i] for i in range(len(names))],
    }
    
    # Tạo DataFrame
    df1 = pd.DataFrame(dataLR)
    df2 = pd.DataFrame(dataMLP)
    df3 = pd.DataFrame(dataLSTM)

    # Tạo đối tượng Excel trong bộ nhớ
    output = io.BytesIO()


    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name='LR', index=False)
        df2.to_excel(writer, sheet_name='MLP', index=False)
        df3.to_excel(writer, sheet_name='LSTM', index=False)
    
    # Tự động điều chỉnh độ rộng của các cột cho từng sheet
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = length + 2  # Thêm khoảng cách để nội dung thoáng hơn

    output.seek(0)

    # Trả về file Excel để tải xuống
    return send_file(output, as_attachment=True, download_name='Dự đoán điểm Trung Bình lớp 12.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')



@app.route("/predict_excel2", methods=['GET', 'POST'])
def predict_excel2():
    log_access('/predict_excel2')
    if request.method == "POST":
        try:
            excel_file = request.files['excel_file']
            df = pd.read_excel(excel_file, header=None) 
            names = df.iloc[:, 0].tolist() 
            df = df.drop(columns=0) 
            df = df.iloc[:, :55]
            pred_arg_arr = df.to_numpy()
            def load_and_predict(model_path, model_type):
                if model_type == 'keras':
                    model = load_model(f"Models/{model_path}")  # Load model Keras
                else:
                    with open(f"Models/{model_path}", 'rb') as model_file:
                        model = joblib.load(model_file)  # Load model với joblib

                if model_type == 'keras' and "LSTM" in model_path:  # Kiểm tra xem đây có phải là model LSTM không
                    return model.predict(pred_arg_arr.reshape(pred_arg_arr.shape[0], 1, pred_arg_arr.shape[1]))
                else:
                    return model.predict(pred_arg_arr)

            # Kiểm tra checkbox TN (Tự nhiên) hay XH (Xã hội)
            nature_checked = True if 'nature' in request.form else False
            subjects = {}

            with ThreadPoolExecutor() as executor:
                if nature_checked:
                    print("call TN")
                    future1 = executor.submit(load_and_predict, "LR_TN_TN.pkl", model_type='joblib')
                    future2 = executor.submit(load_and_predict, "MLP_TN_TN.keras", model_type='keras')
                    future3 = executor.submit(load_and_predict, "LSTM_TN_TN.keras", model_type='keras')
                    subjects = {'mon1': 'Điểm Lý', 'mon1_1': 'Physics', 'mon2': 'Điểm Hóa', 'mon2_1': 'Chemistry', 'mon3': 'Điểm Sinh', 'mon3_1': 'Biology', 'type': 'TN'}
                else:
                    print("call XH")
                    future1 = executor.submit(load_and_predict, "LR_TN_XH.pkl", model_type='joblib')
                    future2 = executor.submit(load_and_predict, "MLP_TN_XH.keras", model_type='keras')
                    future3 = executor.submit(load_and_predict, "LSTM_TN_XH.keras", model_type='keras')
                    subjects = {'mon1': 'Điểm Sử', 'mon1_1': 'History', 'mon2': 'Điểm Địa', 'mon2_1': 'Geography', 'mon3': 'Điểm GDCD', 'mon3_1': 'Civic Education', 'type': 'XH'}

                prediction1 = future1.result()
                prediction2 = future2.result()
                prediction3 = future3.result()

                pred_df1 = predModel2(prediction1)
                pred_df2 = predModel2(prediction2)
                pred_df3 = predModel2(prediction3)

                pred_df1 = pred_df1.map(lambda x: min(round(x, 1), 10))
                pred_df2 = pred_df2.map(lambda x: min(round(x, 1), 10))
                pred_df3 = pred_df3.map(lambda x: min(round(x, 1), 10))


                    

        except Exception as e:
            print(f"Lỗi khi tải mô hình: {e}")
            return {"error": str(e)}

    return render_template('Page4.html',
    kq_Model1=pred_df1.to_dict(orient='list') if pred_df1 is not None else {},
    kq_Model2=pred_df2.to_dict(orient='list') if pred_df2 is not None else {},
    kq_Model3=pred_df3.to_dict(orient='list') if pred_df3 is not None else {},
    subjects=subjects,
    names=names
)

@app.route('/download_excel2', methods=['POST'])
def download_excel2():
    names = request.form.getlist('names')
    kq_Model1 = request.form.getlist('kq_Model1')
    kq_Model2 = request.form.getlist('kq_Model2')
    kq_Model3 = request.form.getlist('kq_Model3')
    subjects = request.form.getlist('subjects')

    names = json.loads(names[0].replace("'", '"'))
    subjects = json.loads(subjects[0].replace("'", '"'))
    kq_Model1 = json.loads(kq_Model1[0].replace("'", '"'))
    kq_Model2 = json.loads(kq_Model2[0].replace("'", '"'))
    kq_Model3 = json.loads(kq_Model3[0].replace("'", '"'))

    dataLR = {
        "Tên": names,

        # Kết quả từ LR model
        "Điểm Toán": [kq_Model1['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn": [kq_Model1['Văn_1'][i] for i in range(len(names))],
        subjects.get('mon1'): [kq_Model1['Lý_1'][i] for i in range(len(names))],
        subjects.get('mon2'): [kq_Model1['Hóa_1'][i] for i in range(len(names))],
        subjects.get('mon3'): [kq_Model1['Sinh_1'][i] for i in range(len(names))],
        "Điểm Anh": [kq_Model1['Anh_1'][i] for i in range(len(names))],
    }

    dataMLP = {
        "Tên": names,
        
        # Kết quả từ MLP model
        "Điểm Toán": [kq_Model2['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn": [kq_Model2['Văn_1'][i] for i in range(len(names))],
        subjects.get('mon1'): [kq_Model2['Lý_1'][i] for i in range(len(names))],
        subjects.get('mon2'): [kq_Model2['Hóa_1'][i] for i in range(len(names))],
        subjects.get('mon3'): [kq_Model2['Sinh_1'][i] for i in range(len(names))],
        "Điểm Anh": [kq_Model2['Anh_1'][i] for i in range(len(names))],
    }

    dataLSTM = {
        "Tên": names,

        # Kết quả từ LSTM model
        "Điểm Toán": [kq_Model3['Toán_1'][i] for i in range(len(names))],
        "Điểm Văn": [kq_Model3['Văn_1'][i] for i in range(len(names))],
        subjects.get('mon1'): [kq_Model3['Lý_1'][i] for i in range(len(names))],
        subjects.get('mon2'): [kq_Model3['Hóa_1'][i] for i in range(len(names))],
        subjects.get('mon3'): [kq_Model3['Sinh_1'][i] for i in range(len(names))],
        "Điểm Anh": [kq_Model3['Anh_1'][i] for i in range(len(names))]
    }
    
    # Tạo DataFrame
    df1 = pd.DataFrame(dataLR)
    df2 = pd.DataFrame(dataMLP)
    df3 = pd.DataFrame(dataLSTM)

    # Tạo đối tượng Excel trong bộ nhớ
    output = io.BytesIO()


    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df1.to_excel(writer, sheet_name='LR', index=False)
        df2.to_excel(writer, sheet_name='MLP', index=False)
        df3.to_excel(writer, sheet_name='LSTM', index=False)
    
    # Tự động điều chỉnh độ rộng của các cột cho từng sheet
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            for column_cells in worksheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                column_letter = get_column_letter(column_cells[0].column)
                worksheet.column_dimensions[column_letter].width = length + 2  # Thêm khoảng cách để nội dung thoáng hơn

    output.seek(0)

    # Trả về file Excel để tải xuống
    return send_file(output, as_attachment=True, download_name='Dự đoán điểm thi Tốt Nghiệp.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


def log_access(route):
    conn = sqlite3.connect('data/your_database.db')
    cur = conn.cursor()

    # Kiểm tra xem đã có bản ghi nào cho ngày hôm nay chưa
    cur.execute('''
        SELECT * FROM countTimeSummit WHERE date_predict = ? AND route = ?
    ''', (date.today(), route))

    result = cur.fetchone()

    if result:
        # Nếu đã có bản ghi cho ngày hôm nay
        if route in ['/predict_excel1', '/predict_excel2']:
            cur.execute('''
                UPDATE countTimeSummit SET count_predict_excel = count_predict_excel + 1 WHERE date_predict = ? AND route = ?
            ''', (date.today(), route))
        elif route in ['/predict1', '/predict2']:
            cur.execute('''
                UPDATE countTimeSummit SET count_predict = count_predict + 1 WHERE date_predict = ? AND route = ?
            ''', (date.today(), route))
    else:
        # Nếu chưa có bản ghi nào cho ngày hôm nay, thêm bản ghi mới
        cur.execute('''
            INSERT INTO countTimeSummit (date_predict, route) 
            VALUES (?, ?)
        ''', (date.today(), route))

    conn.commit()
    conn.close()


if __name__ == "__main__":
    app.run(host="0.0.0.0")
